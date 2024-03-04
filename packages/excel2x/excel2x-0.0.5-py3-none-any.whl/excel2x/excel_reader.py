#!/usr/bin/env python3
# -*- coding: utf-8 -*
"""Excel reader
"""
from pathlib import Path
import openpyxl
import csv

# 支持的数据类型 String, Number, StringArray, NumberArray


def read_book_values(input_file: Path, sheet_index=0) -> list[list[str]]:
    if input_file.suffix == ".xlsx":
        return read_excel_values(input_file, sheet_index)
    if input_file.suffix == ".csv":
        return read_csv_values(input_file)


def read_excel_values(input_file, sheet_index=0) -> list[list[str]]:
    """read_excel_values
    Args:
        input_file (str): 输入 Excel 文件路径
        sheet_index (int): 工作簿序号. Defaults to 0.
        config (ExcelConfig): Excel 配置. Defaults to None.
    Raises:
        Exception: Excel 数据异常
    """
    input_file = Path(input_file)
    workbook = openpyxl.load_workbook(input_file, read_only=True, data_only=True, rich_text=False)
    sheet = workbook[workbook.sheetnames[sheet_index]]
    max_col = sheet.max_column
    max_row = sheet.max_row
    values = []
    i = 0
    for row in sheet.values:
        row_list = []
        for j, value in enumerate(row):
            if i == 0 and _is_empty_str(value):
                max_col = j
            if j == 0 and _is_empty_str(value):
                max_row = i
            if i >= max_row or j >= max_col:
                break
            row_list.append(_strip_str(value))
        if i >= max_row:
            break
        values.append(row_list)
        i += 1
    workbook.close()
    if max_row < 3:
        raise ValueError(f"ExcelError 数据最少不少于 3 行！{input_file.name}")
    return values


def read_csv_values(input_file: str, delimiter=",") -> list[list[str]]:
    _csv_data = []
    _column_count = 0
    with open(input_file, "r", encoding="utf-8") as f:
        _reader = csv.reader(f, delimiter=delimiter)
        for _row in _reader:
            _length = len(_row)
            if _length == 0 or _is_empty_str(_row[0]):
                break
            if _column_count == 0:
                _column_count = _length
            if _length < _column_count:
                _row += [""] * (_column_count - _length)
            elif _length > _column_count:
                _row = _row[:_column_count]
            _csv_data.append([_strip_str(x) for x in _row])
    return _csv_data


def _is_empty_str(value):
    return value is None or len(str(value).strip()) == 0


def _strip_str(value):
    if value is None:
        return ""
    text = str(value).strip()
    # 特殊空格和制表符替换为普通空格 \u00A0=不间断空格, \u200F 是从右到左标记符
    for space in ["\t", "\v", "\f", "\u3000", "\u00A0", "\u200F"]:
        text = text.replace(space, " ")
    # 移除不可见空格
    text = text.replace("\u200B", "")
    # 替换换行符
    text = text.replace("\r", "\n")
    text = text.replace("\n\n", "\n")
    text = text.replace("\n", "\\n")
    text = text.replace('"', '\\"')
    # 法语中叹号前面的空格替换为不间断空格, 防止标点换行到行首
    text = text.replace(" !", "\u00A0!")
    text = text.replace(" .", "\u00A0.")
    return text


def convert_value(value, value_type, key=None, line=None):
    """convert_value"""
    text = str(value).strip()
    if value_type.lower() == "string":
        if text.count('"') % 2 == 1:
            print("\033[1;31m" + f"[ValueError] 包含奇数个双引号! line:{line}, key:{key}, type:{value_type}, value:{value}" + "\033[0m")
        return text
    elif value_type.lower() == "number":
        try:
            float_num = float(text)
            int_num = int(float_num)
            if abs(float_num - int_num) < 0.0001:
                return int_num
            return float_num
        except ValueError:
            if text.lower() == "true":
                return 1
            if text.lower() == "false":
                return 0
    elif value_type.lower().endswith("array"):
        type2 = value_type.replace("array", "").replace("Array", "")
        array = []
        for value2 in text.split(","):
            array.append(convert_value(value2, type2, key, line))
        return array
    raise ValueError(f"[TypeError] 数据类型不匹配！line:{line}, key:{key}, type:{value_type}, value:{value}")
