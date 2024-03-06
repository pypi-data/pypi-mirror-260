import typing

from openpyxl.cell import Cell

from excel_models.typing import TTable, TColumnDef, CellValue
from ._base import BaseExcelColumn


class ExcelColumn(BaseExcelColumn):
    def __init__(
            self,
            table: TTable,
            column_def: TColumnDef,
            col_num: int,
            *,
            concrete: bool = True,
    ):
        super().__init__(table, column_def)
        self.col_num = col_num
        self.concrete = concrete

    def occupies(self, col_num: int) -> bool:
        if not self.concrete:
            return False
        return col_num == self.col_num

    def cell(self, row_num: int) -> Cell:
        return self.table.cell(row_num, self.col_num)

    @property
    def cells(self) -> typing.Sequence[Cell]:
        return self.table.col(self.col_num, data_only=True)

    def get_raw(self, row_num: int) -> CellValue:
        return self.cell(row_num).value

    def set_raw(self, row_num: int, raw: CellValue) -> None:
        self.cell(row_num).value = raw

    def delete_raw(self, row_num: int) -> None:
        self.cell(row_num).value = None
