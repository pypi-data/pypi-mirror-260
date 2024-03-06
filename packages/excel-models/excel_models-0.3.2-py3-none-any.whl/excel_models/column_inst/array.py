import typing
from functools import cached_property

from openpyxl.cell import Cell
from returns import returns

from excel_models.typing import TTable, TColumnDef
from ._base import BaseExcelColumn


class ExcelColumnArray(BaseExcelColumn):
    def __init__(
            self,
            table: TTable,
            column_def: TColumnDef,
            col_num: int,
            width: int,
    ):
        super().__init__(table, column_def)
        self.col_num = col_num
        self.width = width

    @cached_property
    def col_nums(self) -> typing.Sequence[int]:
        return tuple(range(self.col_num, self.col_num + self.width))

    def occupies(self, col_num: int) -> bool:
        return col_num in self.col_nums

    @returns(tuple)
    def cell(self, row_num: int) -> typing.Sequence[Cell]:
        for col_num in self.col_nums:
            yield self.table.cell(row_num, col_num)

    @property
    @returns(list)
    def cells(self) -> typing.Sequence[typing.Sequence[Cell]]:
        for row in self.table.data_rows:
            yield self.cell(row)

    @returns(tuple)
    def get_raw(self, row_num: int) -> typing.Sequence:
        for cell in self.cell(row_num):
            yield cell.value

    def set_raw(self, row_num: int, raw: typing.Sequence) -> None:
        for cell, raw_value in zip(self.cell(row_num), raw, strict=True):
            cell.value = raw_value

    def delete_raw(self, row_num: int) -> None:
        for cell in self.cell(row_num):
            cell.value = None
