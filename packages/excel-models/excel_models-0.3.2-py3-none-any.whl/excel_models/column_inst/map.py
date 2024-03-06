import typing

from openpyxl.cell import Cell
from returns import returns

from excel_models.typing import TTable, TColumnDef
from ._base import BaseExcelColumn


class ExcelColumnMap(BaseExcelColumn):
    def __init__(
            self,
            table: TTable,
            column_def: TColumnDef,
            col_map: typing.Mapping[typing.Hashable, int],
    ):
        super().__init__(table, column_def)
        self.col_map = col_map

    def occupies(self, col_num: int) -> bool:
        return col_num in self.col_map.values()

    @returns(dict)
    def cell(self, row_num: int) -> typing.Mapping[typing.Hashable, Cell]:
        for k, col_num in self.col_map.items():
            yield k, self.table.cell(row_num, col_num)

    @property
    @returns(list)
    def cells(self) -> typing.Sequence[typing.Mapping[typing.Hashable, Cell]]:
        for row in self.table.data_rows:
            yield self.cell(row)

    @returns(dict)
    def get_raw(self, row_num: int) -> typing.Mapping:
        for k, cell in self.cell(row_num).items():
            yield k, cell.value

    def set_raw(self, row_num: int, raw: typing.Mapping) -> None:
        cells = self.cell(row_num)
        for k, v in raw.items():
            cells[k].value = v
        for k, cell in cells.items():
            if k in raw:
                continue
            cell.value = None

    def delete_raw(self, row_num: int) -> None:
        for cell in self.cell(row_num).values():
            cell.value = None
