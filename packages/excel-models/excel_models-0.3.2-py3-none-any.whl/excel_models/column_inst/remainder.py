import typing
from itertools import zip_longest

from openpyxl.cell import Cell
from returns import returns

from excel_models.typing import TTable, TColumnDef
from excel_models.utils.collections import rstrip_none
from ._base import BaseExcelColumn


class ExcelColumnRemainder(BaseExcelColumn):
    def __init__(
            self,
            table: TTable,
            column_def: TColumnDef,
            col_num: int,
    ):
        super().__init__(table, column_def)
        self.col_num = col_num

    def occupies(self, col_num: int) -> bool:
        return col_num >= self.col_num

    def cell(self, row_num: int) -> typing.Sequence[Cell]:
        return self.table.row(row_num, min_col=self.col_num)

    @property
    @returns(list)
    def cells(self) -> typing.Sequence[typing.Sequence[Cell]]:
        for row in self.table.data_rows:
            yield self.cell(row)

    @returns(tuple)
    @returns(rstrip_none)
    def get_raw(self, row_num: int) -> typing.Sequence:
        for cell in self.cell(row_num):
            yield cell.value

    def set_raw(self, row_num: int, raw: typing.Sequence) -> None:
        for i, (cell, raw_value) in enumerate(zip_longest(self.cell(row_num), raw, fillvalue=None)):
            if cell is None:
                # if raw is longer than self.cell(row_num), we need to extend the spreadsheet to the right
                cell = self.table.cell(row_num, self.col_num + i)
            cell.value = raw_value

    def delete_raw(self, row_num: int) -> None:
        for cell in self.cell(row_num):
            cell.value = None
