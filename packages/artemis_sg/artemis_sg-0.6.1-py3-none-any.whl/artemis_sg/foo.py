import math
from copy import copy

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder

from artemis_sg.config import CFG


def get_sheet_keys(ws):
    for row in ws.values:
        sheet_keys = [x.upper() if isinstance(x, str) else x for x in row]
        break
    return sheet_keys

def shift_col(ws, col_key, target_idx):
    ws.insert_cols(target_idx)
    sheet_keys = get_sheet_keys(ws)
    sheet_key_idx = sheet_keys.index(col_key) + 1  # for openpyxl
    sheet_key_idx_ltr = get_column_letter(sheet_key_idx)
    col_delta = target_idx - sheet_key_idx
    ws.move_range(f"{sheet_key_idx_ltr}1:{sheet_key_idx_ltr}{ws.max_row}",
                  rows=0, cols=col_delta)
    ws.delete_cols(sheet_key_idx)

def copy_cell_style(ws, style_src_cell, target_cell):
    if style_src_cell.has_style:
        ws[target_cell].font = copy(style_src_cell.font)
        ws[target_cell].border = copy(style_src_cell.border)
        ws[target_cell].fill = copy(style_src_cell.fill)
        ws[target_cell].number_format = copy(style_src_cell.number_format)
        ws[target_cell].protection = copy(style_src_cell.protection)
        ws[target_cell].alignment = copy(style_src_cell.alignment)

def create_col(ws, col_key, target_idx, style_src_cell=None):
    ws.insert_cols(target_idx)
    col_header = f"{get_column_letter(target_idx)}1"
    ws[col_header] = col_key
    if style_src_cell:
        copy_cell_style(ws, style_src_cell, col_header)

def sequence_worksheet(ws, col_order):
    sheet_keys = get_sheet_keys(ws)
    for i, key_name in enumerate(col_order):
        order_idx = i + 1  # for openpyxl
        get_column_letter(order_idx)
        if key_name == "ISBN":
            key_name = isbn_key  # noqa: PLW2901
        if key_name in sheet_keys:
            shift_col(ws, key_name, order_idx)
        else:
            create_col(ws, key_name, order_idx)

def size_sheet_cols(ws):
    dim_holder = DimensionHolder(worksheet=ws)
    sheet_keys = get_sheet_keys(ws)
    for i, key_name in enumerate(sheet_keys):
        col_idx = i + 1  # for openpyxl
        col_idx_ltr = get_column_letter(col_idx)
        width = (
            max(len(str(cell.value)) for cell in ws[col_idx_ltr])
            * CFG["asg"]["spreadsheet"]["sheet_image"]["col_buffer"]
        )
        if width > CFG["asg"]["spreadsheet"]["sheet_image"]["max_col_width"]:
            width = CFG["asg"]["spreadsheet"]["sheet_image"]["max_col_width"]
        dim_holder[col_idx_ltr] = ColumnDimension(ws, index=col_idx_ltr, width=width)
        if key_name == isbn_key:
            dim_holder[col_idx_ltr] = ColumnDimension(
                ws,
                index=col_idx_ltr,
                width=math.ceil(
                    CFG["asg"]["spreadsheet"]["sheet_image"]["isbn_col_width"]
                    * CFG["asg"]["spreadsheet"]["sheet_image"]["col_buffer"]
                ),
            )
        if key_name == "IMAGE":
            dim_holder[col_idx_ltr] = ColumnDimension(
                ws,
                index=col_idx_ltr,
                width=CFG["asg"]["spreadsheet"]["sheet_image"]["image_col_width"]
            )

    ws.column_dimensions = dim_holder


isbn_key = "ISBN-13"
col_order = CFG["asg"]["spreadsheet"]["sheet_image"]["col_order"]
workbook = "/home/john/Downloads/foo.xlsx"

wb = load_workbook(workbook)
ws = wb.worksheets[0]

sequence_worksheet(ws, col_order)
size_sheet_cols(ws)
wb.save(workbook)
